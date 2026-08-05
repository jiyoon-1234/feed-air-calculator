from __future__ import annotations

import csv
import io
import json
import mimetypes
import os
import re
import sys
import webbrowser
from email import policy
from email.parser import BytesParser
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse


ROOT = Path(__file__).resolve().parent
DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 8766


class FeedAirHandler(SimpleHTTPRequestHandler):
    server_version = "FeedAirCalculator/1.0"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def log_message(self, fmt: str, *args) -> None:
        sys.stdout.write("%s - %s\n" % (self.address_string(), fmt % args))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/parse" and parse_qs(parsed.query).get("template"):
            return self._send_template()
        if self.path in {"/", ""}:
            self.path = "/feed-air-calculator.html"
        return super().do_GET()

    def do_POST(self) -> None:
        if self.path != "/api/parse":
            self.send_error(404, "Not found")
            return
        try:
            filename, data = self._read_uploaded_file()
            rows = parse_uploaded_file(filename, data)
            self._send_json({"ok": True, "filename": filename, "rows": rows})
        except Exception as exc:
            self._send_json({"ok": False, "error": str(exc)}, status=400)

    def _read_uploaded_file(self) -> tuple[str, bytes]:
        content_type = self.headers.get("Content-Type", "")
        content_length = int(self.headers.get("Content-Length", "0") or 0)
        if content_length <= 0:
            raise ValueError("No uploaded file was received.")
        body = self.rfile.read(content_length)

        raw = (
            f"Content-Type: {content_type}\r\n"
            "MIME-Version: 1.0\r\n\r\n"
        ).encode("utf-8") + body
        message = BytesParser(policy=policy.default).parsebytes(raw)
        if not message.is_multipart():
            raise ValueError("Upload request must be multipart/form-data.")

        for part in message.iter_parts():
            disposition = part.get_content_disposition()
            if disposition != "form-data":
                continue
            if part.get_param("name", header="content-disposition") != "file":
                continue
            filename = part.get_filename() or "upload"
            payload = part.get_payload(decode=True)
            if not payload:
                raise ValueError("Uploaded file is empty.")
            return filename, payload
        raise ValueError("Could not find the uploaded file field.")

    def _send_json(self, payload: dict, status: int = 200) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_template(self) -> None:
        template_path = ROOT / "Standard_Composition_Table.xlsx"
        if not template_path.exists():
            self.send_error(404, "Template not found")
            return
        data = template_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", 'attachment; filename="Standard_Composition_Table.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def parse_uploaded_file(filename: str, data: bytes) -> list[list[str]]:
    ext = Path(filename).suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return parse_xlsx(data)
    if ext == ".xls":
        raise ValueError("Legacy .xls files are not supported. Save the workbook as .xlsx or .csv.")
    if ext == ".pdf":
        return parse_pdf(data)
    if ext in {".csv", ".txt", ".tsv"}:
        return parse_csv_text(data)
    guessed = mimetypes.guess_type(filename)[0] or ""
    if "csv" in guessed:
        return parse_csv_text(data)
    raise ValueError("Supported formats are .xlsx, .csv, .tsv, and text-based PDF.")


def parse_xlsx(data: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("openpyxl is not installed, so Excel files cannot be parsed.") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        values = [cell_to_text(cell) for cell in row]
        while values and values[-1] == "":
            values.pop()
        if any(values):
            rows.append(values)
    return rows


def parse_csv_text(data: bytes) -> list[list[str]]:
    text = decode_text(data)
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample else ","
    rows = []
    for row in csv.reader(io.StringIO(text), delimiter=delimiter):
        values = [cell.strip() for cell in row]
        if any(values):
            rows.append(values)
    return rows


def parse_pdf(data: bytes) -> list[list[str]]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise ValueError("pdfplumber is not installed, so PDF files cannot be parsed.") from exc

    rows: list[list[str]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    values = [cell_to_text(cell) for cell in row]
                    while values and values[-1] == "":
                        values.pop()
                    if any(values):
                        rows.append(values)
            if tables:
                continue
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            for line in text.splitlines():
                values = [v.strip() for v in re.split(r"\s{2,}", line.strip()) if v.strip()]
                if values:
                    rows.append(values)
    if not rows:
        raise ValueError("No text table was found in the PDF. Scanned PDFs need OCR first.")
    return rows


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main() -> None:
    os.chdir(ROOT)
    host = os.environ.get("HOST")
    port = int(os.environ.get("PORT", DEFAULT_PORT))
    cloud_mode = "PORT" in os.environ
    bind_host = host or ("0.0.0.0" if cloud_mode else DEFAULT_HOST)
    server, actual_port = make_server(bind_host, port, allow_scan=not cloud_mode)
    display_host = "localhost" if bind_host in {"0.0.0.0", "127.0.0.1"} else bind_host
    url = f"http://{display_host}:{actual_port}/"
    print(f"Feed air calculator running at {url}")

    if "--no-browser" not in sys.argv and not cloud_mode:
        try:
            webbrowser.open(url)
        except Exception:
            pass

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
    finally:
        server.server_close()


def make_server(host: str, port: int, allow_scan: bool) -> tuple[ThreadingHTTPServer, int]:
    candidate_ports = range(port, port + 20) if allow_scan else [port]
    for candidate in candidate_ports:
        try:
            return ThreadingHTTPServer((host, candidate), FeedAirHandler), candidate
        except OSError:
            continue
    raise RuntimeError(f"Could not bind to port {port}.")


if __name__ == "__main__":
    main()
